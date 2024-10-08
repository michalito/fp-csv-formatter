import csv
import io
import re
import pandas as pd
import traceback
from openpyxl import load_workbook, Workbook


# Input size map (unchanged)
INPUT_SIZE_MAP = {'XS': 'XSmall', 'SM': 'Small', 'ME': 'Medium', 'LA': 'Large', 'XL': 'X Large'}

# Output size map (new)
OUTPUT_SIZE_MAP = {'XS': 'XSmall', 'SM': 'Small', 'MD': 'Medium', 'LG': 'Large', 'XL': 'XLarge'}

# Reverse map for converting full size to short code
REVERSE_OUTPUT_SIZE_MAP = {v: k for k, v in OUTPUT_SIZE_MAP.items()}


def process_file(file_content, file_type, product_name, product_sku_base, default_price, wholesale_price, consignment_price, cost, weight, brand, gender, suppliers, sheet_name=None):
    try:
        if file_type == 'csv':
            return process_csv(file_content, product_name, product_sku_base, default_price, wholesale_price, consignment_price, cost, weight, brand, gender, suppliers)
        elif file_type == 'xlsx':
            return process_excel(file_content, product_name, product_sku_base, default_price, wholesale_price, consignment_price, cost, weight, brand, gender, suppliers, sheet_name)
        else:
            raise ValueError(f"Unsupported file type: {file_type}")
    except Exception as e:
        print(f"Error in process_file: {str(e)}")
        traceback.print_exc()
        raise
    
def process_csv(file_content, product_name, product_sku_base, default_price, wholesale_price, consignment_price, cost, weight, brand, gender, suppliers):
    reader = csv.DictReader(io.StringIO(file_content.decode('utf-8-sig')))
    return process_data(reader, product_name, product_sku_base, default_price, wholesale_price, consignment_price, cost, weight, brand, gender, suppliers)

def process_excel(file_content, product_name, product_sku_base, default_price, wholesale_price, consignment_price, cost, weight, brand, gender, suppliers, sheet_name=None):
    wb = load_workbook(filename=io.BytesIO(file_content))
    
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in the workbook")
        ws = wb[sheet_name]
    else:
        ws = wb.active
    
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    reader = [dict(zip(header, row)) for row in rows[1:]]
    return process_data(reader, product_name, product_sku_base, default_price, wholesale_price, consignment_price, cost, weight, brand, gender, suppliers)

def process_data(reader, product_name, product_sku_base, default_price, wholesale_price, consignment_price, cost, weight, brand, gender, suppliers):
    try:
        processed_data = {}
        current_product = None
        current_price = None
        
        for row in reader:
            product_sku = row['Product SKU']
            
            # Check if this is a product row or an item row
            if not any(size in product_sku for size in INPUT_SIZE_MAP.keys()):
                # This is a product row
                color = ' '.join(row['Product Name'].split()[1:])
                current_product = product_sku
                current_price = row['Price'].replace('€', '').strip()
                
            if current_product not in processed_data:
                processed_data[current_product] = {
                    'Product': product_name,
                    'Color': color,
                    'Brand': brand,
                    'Gender': gender,
                    'Suppliers': suppliers,
                    'WholesalePrice': wholesale_price,
                    'ConsignmentPrice': consignment_price,
                    'Cost': cost,
                    'Weight': weight,
                    'Items': {}
                }
            
            else:
                # This is an item row
                size_identifier = product_sku.split('-')[-1]
                full_size = re.search(r'\[S\]Size=(.*?)(?=\s|$)', row['Product Name']).group(1)
                input_size = INPUT_SIZE_MAP.get(size_identifier, full_size.split()[0])
                
                # Convert input size to output size
                output_size_identifier = REVERSE_OUTPUT_SIZE_MAP.get(input_size, size_identifier)
                output_full_size = OUTPUT_SIZE_MAP.get(output_size_identifier, input_size)
                
                color_identifier = row['MPN'][-3:]
                item_sku = f"{product_sku_base}-{color_identifier}-{output_size_identifier}"
                
                processed_data[current_product]['Items'][item_sku] = {
                    'Size': output_size_identifier,
                    'FullSize': output_full_size,
                    'Stock': row['Stock'] or '0',
                    'MPN': row['MPN'],
                    'GTIN': row['GTIN'] or '',
                    'Price': current_price or default_price,
                    'Status': row['Status'] or ''
                }
        
        # Ensure all sizes are present for each product
        all_sizes = list(OUTPUT_SIZE_MAP.keys())
        for product_sku, product_data in processed_data.items():
            color_identifier = next(iter(product_data['Items'].values()))['MPN'][-3:]
            for size in all_sizes:
                item_sku = f"{product_sku_base}-{color_identifier}-{size}"
                if item_sku not in product_data['Items']:
                    product_data['Items'][item_sku] = {
                        'Size': size,
                        'FullSize': OUTPUT_SIZE_MAP[size],
                        'Stock': '0',
                        'MPN': '',
                        'GTIN': '',
                        'Price': current_price or default_price,
                        'Status': ''
                    }
        
        return processed_data
    
    except Exception as e:
        raise Exception(f"Error processing data: {str(e)}")
    
def convert_to_odoo(file_content, file_type, primary_category='', secondary_category='', tertiary_category=''):
    if file_type == 'csv':
        input_file = io.StringIO(file_content.decode('utf-8-sig'))
        reader = csv.DictReader(input_file)
    elif file_type == 'xlsx':
        wb = load_workbook(filename=io.BytesIO(file_content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        header = rows[0]
        reader = [dict(zip(header, row)) for row in rows[1:]]
    else:
        raise ValueError("Unsupported file type")
    
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=[
        'External_ID', 'base_sku', 'Internal Reference', 'Name', 'Product Category (External_ID)',
        'Barcode', 'Supplier Product Code', 'Published', 'Color', 'Size', 'Sales Price', 'Wholesale Price', 'Consignment Price', 'Cost',
        'Weight', 'Package Length (cm)', 'Package Width (cm)', 'Package Height (cm)', 'Brand',
        'Gender', 'Suppliers', 'Primary Supplier', 'Description'
    ])
    writer.writeheader()

    # Construct the category external ID
    category_external_id = f"category_{primary_category.lower()}"
    if secondary_category:
        category_external_id += f"_{secondary_category.lower()}"
    if tertiary_category:
        category_external_id += f"_{tertiary_category.lower()}"

    for row in reader:
        if int(row.get('Stock', 0)) == 0:
            continue

        product_name = row.get('Product', '')
        color = row.get('Color', '')
        size = row.get('Size', '')
        sku = row.get('Item SKU', '')
        price = row.get('Price', '')
        barcode = row.get('GTIN', '')
        mpn = row.get('MPN', '')
        status = row.get('Status', '')
        brand = row.get('Brand', '')
        gender = row.get('Gender', '')
        suppliers = row.get('Suppliers', '')

        base_sku = sku.rsplit('-', 2)[0] if sku else ''
        external_id = f"product_{sku.replace('-', '_')}" if sku else ''

        new_row = {
            'External_ID': external_id,
            'base_sku': base_sku,
            'Internal Reference': sku,
            'Name': f"{product_name} - {color} ({size})",
            'Product Category (External_ID)': category_external_id,
            'Barcode': barcode,
            'Supplier Product Code': mpn,
            'Published': '1' if status and status.lower() == 'active' else '0',
            'Color': color,
            'Size': size,
            'Sales Price': price,
            'Wholesale Price': row.get('Wholesale Price', ''),
            'Consignment Price': row.get('Consignment Price', ''),
            'Cost': row.get('Cost', ''),
            'Weight': row.get('Weight', ''),
            'Package Length (cm)': '',
            'Package Width (cm)': '',
            'Package Height (cm)': '',
            'Brand': brand,
            'Gender': gender,
            'Suppliers': suppliers,
            'Primary Supplier': suppliers,
            'Description': ''
        }
        writer.writerow(new_row)

    return output.getvalue()

def convert_to_odoo_xlsx(file_content, file_type, primary_category='', secondary_category='', tertiary_category=''):
    if file_type == 'csv':
        input_file = io.StringIO(file_content.decode('utf-8-sig'))
        reader = csv.DictReader(input_file)
    elif file_type == 'xlsx':
        wb = load_workbook(filename=io.BytesIO(file_content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        header = rows[0]
        reader = [dict(zip(header, row)) for row in rows[1:]]
    else:
        raise ValueError("Unsupported file type")
    
    output_wb = Workbook()
    output_ws = output_wb.active
    output_ws.title = "Odoo Import"
    
    headers = [
        'External_ID', 'base_sku', 'Internal Reference', 'Name', 'Product Category (External_ID)',
        'Barcode', 'Supplier Product Code', 'Published', 'Color', 'Size', 'Sales Price', 'Wholesale Price', 'Consignment Price', 'Cost',
        'Weight', 'Package Length (cm)', 'Package Width (cm)', 'Package Height (cm)', 'Brand',
        'Gender', 'Suppliers', 'Primary Supplier', 'Description'
    ]
    output_ws.append(headers)

    # Construct the category external ID
    category_external_id = f"category_{primary_category.lower()}"
    if secondary_category:
        category_external_id += f"_{secondary_category.lower()}"
    if tertiary_category:
        category_external_id += f"_{tertiary_category.lower()}"

    for row in reader:
        if int(row.get('Stock', 0)) == 0:
            continue

        product_name = row.get('Product', '')
        color = row.get('Color', '')
        size = row.get('Size', '')
        sku = row.get('Item SKU', '')
        price = row.get('Price', '')
        barcode = row.get('GTIN', '')
        mpn = row.get('MPN', '')
        status = row.get('Status', '')
        brand = row.get('Brand', '')
        gender = row.get('Gender', '')
        suppliers = row.get('Suppliers', '')

        base_sku = sku.rsplit('-', 2)[0] if sku else ''
        external_id = f"product_{sku.replace('-', '_')}" if sku else ''

        new_row = [
            external_id,
            base_sku,
            sku,
            f"{product_name} - {color} ({size})",
            category_external_id,
            barcode,
            mpn,
            '1' if status and status.lower() == 'active' else '0',
            color,
            size,
            price,
            row.get('Wholesale Price', ''),
            row.get('Consignment Price', ''),
            row.get('Cost', ''),
            row.get('Weight', ''),
            '',  # Package Length (cm)
            '',  # Package Width (cm)
            '',  # Package Height (cm)
            brand,
            gender,
            suppliers,
            suppliers,  # Primary Supplier
            ''  # Description
        ]
        output_ws.append(new_row)

    output = io.BytesIO()
    output_wb.save(output)
    output.seek(0)
    return output

def convert_to_odoo_xlsx(file_content, file_type, primary_category='', secondary_category='', tertiary_category=''):
    if file_type == 'csv':
        input_file = io.StringIO(file_content.decode('utf-8-sig'))
        reader = csv.DictReader(input_file)
    elif file_type == 'xlsx':
        wb = load_workbook(filename=io.BytesIO(file_content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        header = rows[0]
        reader = [dict(zip(header, row)) for row in rows[1:]]
    else:
        raise ValueError("Unsupported file type")
    
    output_wb = Workbook()
    output_ws = output_wb.active
    output_ws.title = "Odoo Import"
    
    headers = [
        'External_ID', 'base_sku', 'Internal Reference', 'Name', 'Product Category (External_ID)',
        'Barcode', 'Supplier Product Code', 'Published', 'Color', 'Size', 'Sales Price', 'Wholesale Price', 'Consignment Price', 'Cost',
        'Weight', 'Package Length (cm)', 'Package Width (cm)', 'Package Height (cm)', 'Brand',
        'Gender', 'Suppliers', 'Primary Supplier', 'Description'
    ]
    output_ws.append(headers)

    # Construct the category external ID
    category_external_id = f"category_{primary_category.lower()}"
    if secondary_category:
        category_external_id += f"_{secondary_category.lower()}"
    if tertiary_category:
        category_external_id += f"_{tertiary_category.lower()}"

    for row in reader:
        if int(row.get('Stock', 0)) == 0:
            continue

        product_name = row.get('Product', '')
        color = row.get('Color', '')
        size = row.get('Size', '')
        sku = row.get('Item SKU', '')
        price = row.get('Price', '')
        barcode = row.get('GTIN', '')
        mpn = row.get('MPN', '')
        status = row.get('Status', '')
        brand = row.get('Brand', '')
        gender = row.get('Gender', '')
        suppliers = row.get('Suppliers', '')

        base_sku = sku.rsplit('-', 2)[0] if sku else ''
        external_id = f"product_{sku.replace('-', '_')}" if sku else ''

        new_row = [
            external_id,
            base_sku,
            sku,
            f"{product_name} - {color} ({size})",
            category_external_id,
            barcode,
            mpn,
            '1' if status and status.lower() == 'active' else '0',
            color,
            size,
            price,
            row.get('Wholesale Price', ''),
            row.get('Consignment Price', ''),
            row.get('Cost', ''),
            row.get('Weight', ''),
            '',  # Package Length (cm)
            '',  # Package Width (cm)
            '',  # Package Height (cm)
            brand,
            gender,
            suppliers,
            suppliers,  # Primary Supplier
            ''  # Description
        ]
        output_ws.append(new_row)

    output = io.BytesIO()
    output_wb.save(output)
    output.seek(0)
    return output

def generate_csv(processed_data):
    try:
        output = io.StringIO()
        writer = csv.writer(output)
        
        writer.writerow(['Product', 'Item', 'Item SKU', 'Color', 'Size', 'Stock', 'MPN', 'GTIN', 'Price', 'Wholesale Price', 'Consignment Price', 'Cost', 'Weight', 'Status', 'Brand', 'Gender', 'Suppliers'])
        
        for product_sku, product_data in processed_data.items():
            product = product_data['Product']
            color = product_data['Color']
            brand = product_data['Brand']
            gender = product_data['Gender']
            suppliers = product_data['Suppliers']
            wholesale_price = product_data['WholesalePrice']
            consignment_price = product_data['ConsignmentPrice']
            cost = product_data['Cost']
            weight = product_data['Weight']
            for item_sku, item_data in product_data['Items'].items():
                item = f"{product} {color} {item_data['Size']}"
                writer.writerow([
                    product,
                    item,
                    item_sku,
                    color,
                    item_data['FullSize'],
                    item_data['Stock'],
                    item_data['MPN'],
                    item_data['GTIN'],
                    item_data['Price'],
                    wholesale_price,
                    consignment_price,
                    cost,
                    weight,
                    item_data['Status'],
                    brand,
                    gender,
                    suppliers
                ])
        
        return output.getvalue()
    except Exception as e:
        raise Exception(f"Error generating CSV: {str(e)}")

def get_excel_sheet_names(file_content):
    wb = load_workbook(filename=io.BytesIO(file_content))
    return wb.sheetnames

def get_initial_product_info(file_content, file_type, sheet_name=None):
    if file_type == 'csv':
        return get_initial_product_info_csv(file_content)
    elif file_type == 'xlsx':
        return get_initial_product_info_excel(file_content, sheet_name)
    else:
        raise ValueError("Unsupported file type")

def get_initial_product_info_csv(file_content):
    try:
        reader = csv.DictReader(io.StringIO(file_content.decode('utf-8-sig')))
        first_row = next(reader)
        product_name = first_row['Product Name'].split()[0]
        product_sku_base = first_row['Product SKU'].split('-')[0]
        return product_name, product_sku_base
    except Exception as e:
        raise Exception(f"Error getting initial product info from CSV: {str(e)}")

def get_initial_product_info_excel(file_content, sheet_name=None):
    try:
        wb = load_workbook(filename=io.BytesIO(file_content))
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in the workbook")
            ws = wb[sheet_name]
        else:
            ws = wb.active
        
        first_row = next(ws.iter_rows(values_only=True))
        header = first_row
        data_row = next(ws.iter_rows(values_only=True))
        first_row_dict = dict(zip(header, data_row))
        
        product_name = first_row_dict['Product Name'].split()[0]
        product_sku_base = first_row_dict['Product SKU'].split('-')[0]
        return product_name, product_sku_base
    except Exception as e:
        raise Exception(f"Error getting initial product info from Excel: {str(e)}")
    
def generate_xlsx(processed_data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Processed Inventory"
    
    headers = ['Product', 'Item', 'Item SKU', 'Color', 'Size', 'Stock', 'MPN', 'GTIN', 'Price', 'Wholesale Price', 'Consignment Price', 'Cost', 'Weight', 'Status', 'Brand', 'Gender', 'Suppliers']
    ws.append(headers)
    
    for product_sku, product_data in processed_data.items():
        product = product_data['Product']
        color = product_data['Color']
        brand = product_data['Brand']
        gender = product_data['Gender']
        suppliers = product_data['Suppliers']
        wholesale_price = product_data['WholesalePrice']
        consignment_price = product_data['ConsignmentPrice']
        cost = product_data['Cost']
        weight = product_data['Weight']
        for item_sku, item_data in product_data['Items'].items():
            item = f"{product} {color} {item_data['Size']}"
            ws.append([
                product,
                item,
                item_sku,
                color,
                item_data['FullSize'],
                item_data['Stock'],
                item_data['MPN'],
                item_data['GTIN'],
                item_data['Price'],
                wholesale_price,
                consignment_price,
                cost,
                weight,
                item_data['Status'],
                brand,
                gender,
                suppliers
            ])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def generate_stock_move(file_content, file_type, location):
    if file_type == 'csv':
        df = pd.read_csv(io.StringIO(file_content.decode('utf-8-sig')))
    elif file_type == 'xlsx':
        df = pd.read_excel(io.BytesIO(file_content))
    else:
        raise ValueError("Unsupported file type")

    stock_move_data = []

    for _, row in df.iterrows():
        if int(row['Stock']) > 0:
            item_sku = row['Item SKU']
            stock_move_data.append({
                'external_id': f"stock_{item_sku.replace('-', '_')}",
                'Product/external_id': f"product_{item_sku.replace('-', '_')}",
                'Product': item_sku,
                'Location': location,
                'Quantity (On Hand)': 0,
                'Counted Quantity': int(row['Stock']),
                'Difference': 0,
                'Scheduled Date': '',
                'Assigned To': 'Administrator'
            })

    return pd.DataFrame(stock_move_data)